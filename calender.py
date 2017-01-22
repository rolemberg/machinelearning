from __future__ import print_function
import httplib2
import os
import dateutil.parser

from apiclient import discovery
import oauth2client
from oauth2client import client
from oauth2client import tools

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

import datetime

try:
    import argparse
    flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
except ImportError:
    flags = None

SCOPES = 'https://www.googleapis.com/auth/calendar.readonly'
CLIENT_SECRET_FILE = 'client_secret.json'
APPLICATION_NAME = 'SAS Python Calendar'


def get_credentials():

    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials')
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir,
                                   'sas-calendar-credential.json')

    store = oauth2client.file.Storage(credential_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES)
        flow.user_agent = APPLICATION_NAME
        if flags:
            credentials = tools.run_flow(flow, store, flags)
        else: # Needed only for compatibility with Python 2.6
            credentials = tools.run(flow, store)
        print('Storing credentials to ' + credential_path)
    return credentials

def main():
    credentials = get_credentials()
    http = credentials.authorize(httplib2.Http())
    service = discovery.build('calendar', 'v3', http=http)

    print("GEARQ - Arquitetura Corporativa Sulamérica\n")
    print("SAS PYTHON CALENDAR EXPORTER\n\n")
    
    sas = {"Pinheiros":['sulamerica.com.br_323531343836383130@resource.calendar.google.com','sulamerica.com.br_31393037333632313738@resource.calendar.google.com','sulamerica.com.br_38393636313434302d343938@resource.calendar.google.com','sulamerica.com.br_2d3337383538343630343138@resource.calendar.google.com','sulamerica.com.br_2d32373231323631382d343432@resource.calendar.google.com','sulamerica.com.br_39383736323437322d333432@resource.calendar.google.com','sulamerica.com.br_3537373433333439323836@resource.calendar.google.com','sulamerica.com.br_3937363031362d323839@resource.calendar.google.com','sulamerica.com.br_35353639313237302d393632@resource.calendar.google.com','sulamerica.com.br_39393536383532333039@resource.calendar.google.com','sulamerica.com.br_2d39323232343739362d383536@resource.calendar.google.com','sulamerica.com.br_2d38363739303637352d373536@resource.calendar.google.com','sulamerica.com.br_39343733313731343634@resource.calendar.google.com','sulamerica.com.br_35393134353937362d32@resource.calendar.google.com','sulamerica.com.br_3734393737363138353336@resource.calendar.google.com','sulamerica.com.br_2d3531323938323530313738@resource.calendar.google.com','sulamerica.com.br_3436333339363931353630@resource.calendar.google.com'],"Matriz":['sulamerica.com.br_2d3437303534343833353536@resource.calendar.google.com','sulamerica.com.br_36323333363234372d363331@resource.calendar.google.com','sulamerica.com.br_3838363932333137353937@resource.calendar.google.com','sulamerica.com.br_37383933383639382d343036@resource.calendar.google.com','sulamerica.com.br_36393834313732392d313038@resource.calendar.google.com','sulamerica.com.br_3731393433393036323037@resource.calendar.google.com','sulamerica.com.br_2d323137393232302d343639@resource.calendar.google.com','sulamerica.com.br_2d32373536303234342d383733@resource.calendar.google.com','sulamerica.com.br_383730333236322d33@resource.calendar.google.com','sulamerica.com.br_2d32383837373937333537@resource.calendar.google.com','sulamerica.com.br_2d3437343330343637343436@resource.calendar.google.com','sulamerica.com.br_313037383334343433@resource.calendar.google.com','sulamerica.com.br_343332383839352d343030@resource.calendar.google.com','sulamerica.com.br_36313133343534302d353234@resource.calendar.google.com','sulamerica.com.br_2d373030393432352d313136@resource.calendar.google.com','sulamerica.com.br_2d37363833393434342d333137@resource.calendar.google.com','sulamerica.com.br_3236383836393838313738@resource.calendar.google.com','sulamerica.com.br_2d36383832323234323239@resource.calendar.google.com','sulamerica.com.br_39393436373035392d383830@resource.calendar.google.com','sulamerica.com.br_3130343838343639323933@resource.calendar.google.com','sulamerica.com.br_2d3137323837323039333936@resource.calendar.google.com','sulamerica.com.br_31313731363439373330@resource.calendar.google.com','sulamerica.com.br_2d35363332363136392d353331@resource.calendar.google.com','sulamerica.com.br_2d31363733363630332d383438@resource.calendar.google.com','sulamerica.com.br_2d3635343631343034323236@resource.calendar.google.com','sulamerica.com.br_2d3537303331383637323839@resource.calendar.google.com','sulamerica.com.br_2d3531343434333936393138@resource.calendar.google.com','sulamerica.com.br_2d32353232313434352d343732@resource.calendar.google.com','sulamerica.com.br_38333134333030382d323330@resource.calendar.google.com','sulamerica.com.br_33373636373932312d363030@resource.calendar.google.com','sulamerica.com.br_2d3535363939343538393935@resource.calendar.google.com','sulamerica.com.br_2d3233383636323934323234@resource.calendar.google.com','sulamerica.com.br_34303333333734313931@resource.calendar.google.com','sulamerica.com.br_33313038313839393937@resource.calendar.google.com','sulamerica.com.br_33393431383134373236@resource.calendar.google.com'],"Morumbi":['sulamerica.com.br_2d36363434343234363037@resource.calendar.google.com','sulamerica.com.br_3938353834313237353637@resource.calendar.google.com','sulamerica.com.br_31343735373836382d353038@resource.calendar.google.com','sulamerica.com.br_38373938383337392d393833@resource.calendar.google.com','sulamerica.com.br_3438353331393531313731@resource.calendar.google.com','sulamerica.com.br_32373731363437312d383937@resource.calendar.google.com','sulamerica.com.br_35363539323336312d363334@resource.calendar.google.com','sulamerica.com.br_31313839363435382d343932@resource.calendar.google.com','sulamerica.com.br_36363231313731352d373537@resource.calendar.google.com','sulamerica.com.br_38373637353834312d343136@resource.calendar.google.com','sulamerica.com.br_2d39363732353739332d323031@resource.calendar.google.com','sulamerica.com.br_2d3832393134303638333031@resource.calendar.google.com','sulamerica.com.br_353631333735382d363233@resource.calendar.google.com','sulamerica.com.br_2d39323238393032352d353630@resource.calendar.google.com','sulamerica.com.br_3632313132343937363531@resource.calendar.google.com','sulamerica.com.br_3634313336383436383432@resource.calendar.google.com','sulamerica.com.br_323134383837333738@resource.calendar.google.com','sulamerica.com.br_37383131383238343530@resource.calendar.google.com','sulamerica.com.br_3939373831363532353033@resource.calendar.google.com','sulamerica.com.br_32323636343734372d313430@resource.calendar.google.com','sulamerica.com.br_2d35313537333436302d393730@resource.calendar.google.com','sulamerica.com.br_2d3639383736363232353135@resource.calendar.google.com','sulamerica.com.br_2d3934303730323439313137@resource.calendar.google.com']}
    eventsResult = service.calendarList().list().execute()
    schedules = eventsResult.get('items', [])

    wb = Workbook()
    dest_filename = 'sas-calendar.xlsx'

    for place in sas:
    
        print("Extraindo eventos para salas de " + place + "\n")

        wx = wb.active
        ws = wb.create_sheet(title=place)
        ws.title = place
        ws.append(["ID","Sala","Assunto","Link","Responsável","Email","DataInício","HoraInício","DataFim","HoraFim","N.Participantes"])

        for room in sas[place]:
            meetings = []
            eventsMax = []
            request = service.events().list(calendarId=room, timeMin='2015-06-01T00:00:00Z',timeMax=datetime.datetime.utcnow().isoformat() + 'Z',orderBy='updated')
            
            while (request != None):
                events = request.execute(http=http)
                eventsMax.append(events)
                request = service.events().list_next(request, events)
            
            location = events['summary']	    
            print(location + "\n")            
            for events in eventsMax:
                meetings = meetings + events['items']

            for meet in meetings:
                if(meet['status'] == 'confirmed'):
                    
                    summary = meet['summary'] if 'summary' in meet else 'Sem título'
                    people = len(meet['attendees']) if 'attendees' in meet else 0
                    link = meet['htmlLink'] if 'htmlLink' in meet else 'indisponível'
                    name = meet['creator']['displayName'] if 'displayName' in meet['creator'] else 'indisponível'
                    start = dateutil.parser.parse(str(meet['start']['dateTime'])) if 'dateTime' in meet['start']  else dateutil.parser.parse(str(meet['start']['date']))
                    end = dateutil.parser.parse(str(meet['end']['dateTime'])) if 'dateTime' in meet['end']  else dateutil.parser.parse(str(meet['end']['date']))
                    
                    if(start=="A"):
                        erro = erro + 1

                    ws.append([str(meet['id']),location.replace(";"," "),summary.replace(";",""),link,name,str(meet['creator']['email']),str(start.strftime("%d/%m/%y")),str(start.strftime("%H:%M:%S")),str(end.strftime("%d/%m/%y")),str(end.strftime("%H:%M:%S")),str(people)])
                    
    print("Salvando XLSX\n\n")
    wb.remove_sheet(wx)
    wb.save(filename = dest_filename)	
    print("Feito.\n")

if __name__ == '__main__':
    main()